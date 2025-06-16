import json
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, GradientFill, Alignment
from openpyxl.utils import get_column_letter
import hashlib

# ==============================================================================
#  STEP 1: DEFINE THE COLOR PALETTES
# ==============================================================================
distinct_100_palette = [
    '#E6194B', '#4363D8', '#FFE119', '#3CB44B', '#F58231', '#911EB4', '#42D4F4', '#F032E6',
    '#BFEF45', '#FABED4', '#469990', '#DCBEFF', '#9A6324', '#FFFAC8', '#800000',
    '#F58282', '#9A132F', '#FFA07A', '#FF7F50', '#DC143C', '#F5A65B', '#B35407',
    '#FFDAB9', '#FF8C00', '#CC5500', '#FFFF80', '#FFD700', '#F0E68C', '#B39E11',
    '#8B4513', '#F5F5DC', '#D2B48C', '#5C3D1B', '#AAFFC3', '#7FFFD4', '#006400',
    '#808000', '#7FFF00', '#228B22', '#AAF0D1', '#008080', '#556B2F', '#80FFFF',
    '#008B8B', '#00FFFF', '#40E0D0', '#20B2AA', '#5F9EA0', '#4682B4', '#87CEEB',
    '#A9D0F5', '#000075', '#0033A0', '#4169E1', '#1E90FF', '#191970', '#4B0082',
    '#6495ED', '#E6E6FA', '#D8BFD8', '#5A007B', '#FF00FF', '#DA70D6', '#D9007E',
    '#9370DB', '#8B008B', '#9966CC', '#6A5ACD', '#404040', '#808080', '#A9A9A9',
    '#C0C0C0', '#DCDCDC', '#F5F5F5', '#ADFF2F', '#FF69B4', '#00BFFF', '#3CB371',
    '#483D8B', '#9400D3', '#B22222', '#B8860B', '#48D1CC', '#9932CC', '#CD853F',
    '#BC8F8F', '#8FBC8F', '#2F4F4F', '#696969', '#CD5C5C', '#778899', '#B0E0E6',
    '#FFC0CB', '#DDA0DD', '#FA8072', '#F4A460', '#98FB98', '#ADD8E6'
]
colorblind_safe_palette = [
    '#377eb8', '#ff7f00', '#4daf4a', '#f781bf', '#a65628',
    '#984ea3', '#999999', '#e41a1c', '#dede00'
]

# ==============================================================================
#  STEP 2: DEFINE THE VISUAL DATA FLOW REPORTER CLASS (CORRECTED)
# ==============================================================================

class VisualDataFlowReporter:
    def __init__(self, palette=None):
        self.search_results = []
        self.filtered_results = []
        self.answers = []
        self.result_connections = {}
        self.source_color_map = {}
        self.color_palette = palette if palette is not None else ['#BDE6F1', '#FDFD96', '#FFB48A']

    def create_content_fingerprint(self, result_dict):
        key_fields = ['title', 'url', 'content', 'description', 'text']
        key_values = [str(result_dict.get(f, '') or '').strip().lower() for f in key_fields]
        combined_content = '|||'.join(key_values)
        return hashlib.md5(combined_content.encode('utf-8')).hexdigest()[:8]

    def load_and_analyze_data(self, json_file_paths):
        for file_path in json_file_paths:
            with open(file_path, 'r', encoding='utf-8') as f: data = json.load(f)
            self._process_search_results(data, file_path)
            self._process_filtered_results(data, file_path)
            self._process_answers(data, file_path)
        self._assign_source_colors()

    # BUG FIX: This method now correctly calculates and adds the 'fingerprint' key.
    def _process_search_results(self, data, file_path):
        for r in data.get('search_results', []):
            fp = self.create_content_fingerprint(r)
            self.search_results.append({**r, 'fingerprint': fp, 'source_file': Path(file_path).name})

    # BUG FIX: This method now correctly adds the 'fingerprint' key.
    def _process_filtered_results(self, data, file_path):
        for r in data.get('filtered_search_results', []):
            fp = self.create_content_fingerprint(r)
            self.result_connections.setdefault(fp, {})['filtered'] = True
            self.filtered_results.append({**r, 'fingerprint': fp, 'source_file': Path(file_path).name})

    def _process_answers(self, data, file_path):
        for r in data.get('answers_to_questions', []):
            fps = [self.create_content_fingerprint(ref) for ref in r.get('referenced_results', [])]
            for fp in fps: self.result_connections.setdefault(fp, {})['used_in_answer'] = True
            self.answers.append({**r, 'referenced_fingerprints': fps, 'source_file': Path(file_path).name})
    
    def _assign_source_colors(self):
        successful_fingerprints = {fp for ans in self.answers for fp in ans['referenced_fingerprints']}
        color_index = 0
        for result in self.search_results:
            # This line now works because 'fingerprint' key is guaranteed to exist.
            if result['fingerprint'] in successful_fingerprints:
                if result['fingerprint'] not in self.source_color_map:
                    if color_index < len(self.color_palette):
                        self.source_color_map[result['fingerprint']] = self.color_palette[color_index]
                        color_index += 1
                    else:
                        self.source_color_map[result['fingerprint']] = self.color_palette[color_index % len(self.color_palette)]
                        color_index += 1

    def create_visual_flow_on_worksheet(self, ws):
        regions = {'search': {'start_col': 1, 'width': 5, 'title': '1. SEARCH RESULTS'},
                   'filtered': {'start_col': 7, 'width': 5, 'title': '2. FILTERED RESULTS'},
                   'answers': {'start_col': 13, 'width': 6, 'title': '3. FINAL ANSWERS'}}
        self._create_region_headers(ws, regions)
        self._populate_regions(ws, regions)
        self._apply_global_formatting(ws)

    def _create_region_headers(self, ws, regions):
        for name, info in regions.items():
            ws.merge_cells(start_row=1, start_column=info['start_col'], end_row=1, end_column=info['start_col'] + info['width'] - 1)
            cell = ws.cell(row=1, column=info['start_col'], value=info['title'])
            cell.font = Font(size=14, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            colors = {'search': "1565C0", 'filtered': "2E7D32", 'answers': "C62828"}
            cell.fill = PatternFill(fill_type="solid", start_color=colors[name])

    # BUG FIX: This method is now simpler and correctly uses the pre-calculated fingerprint.
    def _populate_regions(self, ws, regions):
        max_len = max(len(self.search_results), len(self.filtered_results), len(self.answers))
        headers = {'search': ['File', 'Pos', 'Fingerprint', 'Title', 'Status'],
                   'filtered': ['File', 'Fingerprint', 'Title', 'From Pos', 'Status'],
                   'answers': ['File', 'Question', 'Answer', 'Src Count', 'Sources', 'Value']}
        for region, h_list in headers.items():
            for i, h in enumerate(h_list): ws.cell(row=2, column=regions[region]['start_col'] + i, value=h).font = Font(bold=True)
        for i in range(max_len):
            row = 3 + i
            if i < len(self.search_results):
                r = self.search_results[i]; fp = r['fingerprint']; conn = self.result_connections.get(fp, {})
                color = self.source_color_map.get(fp)
                if color: status = "➡️ Used in Answer"
                elif conn.get('filtered'): status, color = "Filtered", "FFF9C4"
                else: status, color = "Stopped", "FFCDD2"
                self._fill_row(ws, row, regions['search']['start_col'], [r.get('source_file',''), i + 1, fp, r.get('title', '')[:30]+"...", status], PatternFill(start_color=color, fill_type="solid"))
            if i < len(self.filtered_results):
                r = self.filtered_results[i]; fp = r['fingerprint']
                color = self.source_color_map.get(fp)
                if color: status = "➡️ Used in Answer"
                else: status, color = "Unused", "FFF9C4"
                self._fill_row(ws, row, regions['filtered']['start_col'], [r.get('source_file',''), fp, r.get('title', '')[:25]+"...", "n/a", status], PatternFill(start_color=color, fill_type="solid"))
            if i < len(self.answers):
                r = self.answers[i]; source_fps = r['referenced_fingerprints']
                source_colors = [self.source_color_map[fp] for fp in source_fps if fp in self.source_color_map]
                fill = PatternFill(start_color="FFCDD2", fill_type="solid")
                if len(source_colors) == 1: fill = PatternFill(start_color=source_colors[0], fill_type="solid")
                elif len(source_colors) > 1: fill = GradientFill(stop=(source_colors[0], source_colors[1]))
                self._fill_row(ws, row, regions['answers']['start_col'], [r.get('source_file',''), r.get('question', '')[:20]+"...", r.get('answer', '')[:30]+"...", len(source_fps), ", ".join(source_fps)[:30], "High" if len(source_fps) >= 2 else "Low"], fill)

    def _fill_row(self, ws, row, start_col, data, fill_style):
        for i, value in enumerate(data): ws.cell(row=row, column=start_col + i, value=value).fill = fill_style

    def _apply_global_formatting(self, ws):
        for col_letter in ['D','I','N']: ws.column_dimensions[col_letter].width = 35
        for col_letter in ['A','G','M']: ws.column_dimensions[col_letter].width = 15
        for col_letter in ['C','H']: ws.column_dimensions[col_letter].width = 12

# ==============================================================================
#  STEP 3: SETUP DUMMY DATA AND DEFINE ANSWER SETS
# ==============================================================================
def setup_dummy_json_files():
    data1_content = {"search_results": [{"title": "Core AI Concepts", "text": "AI is the simulation..."}],"filtered_search_results": [{"title": "Core AI Concepts", "text": "AI is the simulation..."}],"answers_to_questions": [{"question": "What are fundamentals of AI?", "answer": "The fundamentals of AI...", "referenced_results": [{"title": "Core AI Concepts", "text": "AI is the simulation..."}]}]}
    data2_content = {"search_results": [{"title": "AI in Healthcare", "text": "AI helps..."},{"title": "Robotics in Surgery", "text": "Da Vinci..."}],"filtered_search_results": [{"title": "AI in Healthcare", "text": "AI helps..."},{"title": "Robotics in Surgery", "text": "Da Vinci..."}],"answers_to_questions": [{"question": "AI in medicine?", "answer": "Diagnostics and surgery.", "referenced_results": [{"title": "AI in Healthcare", "text": "AI helps..."},{"title": "Robotics in Surgery", "text": "Da Vinci..."}]}]}
    with open("report_data_1.json", 'w') as f: json.dump(data1_content, f, indent=2)
    with open("report_data_2.json", 'w') as f: json.dump(data2_content, f, indent=2)

# ==============================================================================
#  STEP 4: MAIN SCRIPT TO GENERATE THE MULTI-TAB DASHBOARD
# ==============================================================================
if __name__ == "__main__":
    setup_dummy_json_files()
    
    # --- CONFIGURATION ---
    USE_COLORBLIND_SAFE_PALETTE = False
    ANSWER_SETS = {
        "AI Fundamentals": ["report_data_1.json"],
        "Medical AI Report": ["report_data_2.json"],
        "Full Combined Report": ["report_data_1.json", "report_data_2.json"]
    }
    # ---------------------

    output_excel_file = "Dashboard_Data_Flow_Report.xlsx"
    wb = openpyxl.Workbook()
    
    ws_home = wb.active
    ws_home.title = "Control Panel"
    ws_home['A1'].value = "Data Flow Dashboard"
    ws_home['A1'].font = Font(size=24, bold=True, color="1565C0")
    ws_home['A3'].value = "Select a report to view:"
    ws_home['A3'].font = Font(size=14, italic=True)
    
    for i, set_name in enumerate(ANSWER_SETS.keys()):
        cell = ws_home.cell(row=i + 4, column=1)
        cell.value = set_name
        cell.hyperlink = f"#'{set_name}'!A1"
        cell.font = Font(size=12, color="0000FF", underline="single")

    note_start_row = len(ANSWER_SETS) + 6
    ws_home.cell(row=note_start_row, column=1).value = "Methodology Notes"
    ws_home.cell(row=note_start_row, column=1).font = Font(size=14, bold=True)
    note_cell = ws_home.cell(row=note_start_row + 1, column=1)
    note_cell.value = ("Color Trails: Each successful search result is assigned a unique color from a palette designed for visual distinction. This color 'stains' the data as it flows through the pipeline. Gradient fills in the 'Final Answers' region indicate that the answer was synthesized from multiple, differently-colored sources.")
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_home.merge_cells(start_row=note_start_row + 1, start_column=1, end_row=note_start_row + 4, end_column=6)

    active_palette = colorblind_safe_palette if USE_COLORBLIND_SAFE_PALETTE else distinct_100_palette
    print(f"Using {'Color-Blind Safe' if USE_COLORBLIND_SAFE_PALETTE else 'Full 100-Color'} Palette.")

    for set_name, json_files in ANSWER_SETS.items():
        print(f"--- Generating report for: {set_name} ---")
        ws_report = wb.create_sheet(title=set_name)
        reporter = VisualDataFlowReporter(palette=active_palette)
        reporter.load_and_analyze_data(json_files)
        reporter.create_visual_flow_on_worksheet(ws_report)
        
    print(f"\n✅ Dashboard report created successfully: {output_excel_file}")
    wb.save(output_excel_file)
